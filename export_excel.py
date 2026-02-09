from __future__ import annotations

from io import BytesIO
from typing import Tuple

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def _make_unique_columns(cols) -> list[str]:
    cols = [str(c).strip() if str(c).strip() else "col" for c in cols]
    if len(set(cols)) == len(cols):
        return cols

    seen = {}
    out = []
    for c in cols:
        seen[c] = seen.get(c, 0) + 1
        out.append(c if seen[c] == 1 else f"{c}_{seen[c]}")
    return out


def _autosize_columns(ws, max_width: int = 60, sample_rows: int = 200) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col[:sample_rows]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _format_sheet_as_table(ws, df: pd.DataFrame, table_name: str, style_name: str) -> None:
    """
    - Freeze header
    - Auto-width
    - Table Excel (filtre intégré)
    """
    ws.freeze_panes = "A2"

    nrows, ncols = df.shape
    if ncols == 0:
        return

    last_col = get_column_letter(ncols)
    last_row = nrows + 1  # + header
    table_ref = f"A1:{last_col}{last_row}"

    _autosize_columns(ws)

    # Table Excel uniquement si au moins 1 ligne de données
    if nrows >= 1:
        t = Table(displayName=table_name, ref=table_ref)
        t.tableStyleInfo = TableStyleInfo(
            name=style_name,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(t)


def build_resume(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Produit deux dataframes:
    - résumé par SIREN (ou SIREN demandé si présent)
    - résumé global
    """
    d = df.copy()

    group_col = "SIREN demandé" if "SIREN demandé" in d.columns else ("SIREN" if "SIREN" in d.columns else None)
    if group_col is None:
        return pd.DataFrame(), pd.DataFrame()

    # Normalisation
    if "Siège" in d.columns:
        d["Siège"] = d["Siège"].fillna(False).astype(bool)

    if "État administratif" in d.columns:
        d["Actif?"] = d["État administratif"].astype(str).str.lower().str.contains("actif")
    else:
        d["Actif?"] = pd.NA

    # Résumé par SIREN (colonne finale toujours nommée "SIREN")
    res = (
        d.groupby(group_col)
        .agg(**{
            "Nb SIRET": (("SIRET" if "SIRET" in d.columns else group_col),
                        "nunique" if "SIRET" in d.columns else "size"),
        })
        .reset_index()
        .rename(columns={group_col: "SIREN"})
    )

    # Merge helper (évite de supprimer SIREN quand group_col == "SIREN")
    def merge_metric(res_df: pd.DataFrame, metric_df: pd.DataFrame, metric_col_original: str) -> pd.DataFrame:
        # metric_df contient la clé `group_col` + la métrique renommée
        merged = res_df.merge(metric_df, left_on="SIREN", right_on=group_col, how="left")
        if group_col != "SIREN":
            merged = merged.drop(columns=[group_col])
        return merged

    if "Actif?" in d.columns:
        actifs = (
            d.groupby(group_col)["Actif?"]
            .sum(min_count=1)
            .reset_index()
            .rename(columns={"Actif?": "Nb actifs"})
        )
        res = merge_metric(res, actifs, "Actif?")
        res["Nb actifs"] = res["Nb actifs"].fillna(0).astype(int)
        res["Nb fermés"] = (res["Nb SIRET"] - res["Nb actifs"]).clip(lower=0)

    if "Siège" in d.columns:
        sieges = (
            d.groupby(group_col)["Siège"]
            .sum(min_count=1)
            .reset_index()
            .rename(columns={"Siège": "Nb sièges"})
        )
        res = merge_metric(res, sieges, "Siège")
        res["Nb sièges"] = res["Nb sièges"].fillna(0).astype(int)

    # Résumé global
    global_rows = []
    global_rows.append(("Nb SIREN", d[group_col].nunique()))
    if "SIRET" in d.columns:
        global_rows.append(("Nb SIRET", d["SIRET"].nunique()))
    else:
        global_rows.append(("Nb lignes", len(d)))

    if "Actif?" in d.columns:
        global_rows.append(("Nb actifs", int(d["Actif?"].sum(skipna=True))))
        global_rows.append(("Nb fermés", int((~d["Actif?"]).sum(skipna=True))))

    if "Siège" in d.columns:
        global_rows.append(("Nb sièges", int(d["Siège"].sum(skipna=True))))

    res_global = pd.DataFrame(global_rows, columns=["Indicateur", "Valeur"])

    # Tri lisible
    if "Nb SIRET" in res.columns:
        res = res.sort_values(["Nb SIRET", "SIREN"], ascending=[False, True])

    return res, res_global



def export_sirets_xlsx(df: pd.DataFrame) -> bytes:
    """
    Export XLSX "propre" :
    - Onglet SIRET : Table Excel filtrable
    - Onglet Résumé : global + par SIREN
    """
    d = df.copy()
    d.columns = _make_unique_columns(d.columns)

    resume_par_siren, resume_global = build_resume(d)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Onglet SIRET
        d.to_excel(writer, index=False, sheet_name="SIRET")
        ws_s = writer.book["SIRET"]
        _format_sheet_as_table(ws_s, d, table_name="T_SIRET", style_name="TableStyleMedium9")

        # Onglet Résumé
        start_row = 0
        if not resume_global.empty:
            resume_global.to_excel(writer, index=False, sheet_name="Résumé", startrow=start_row)
            start_row += len(resume_global) + 3

        if not resume_par_siren.empty:
            resume_par_siren.to_excel(writer, index=False, sheet_name="Résumé", startrow=start_row)

        ws_r = writer.book["Résumé"]
        ws_r.freeze_panes = "A2"
        _autosize_columns(ws_r)

        # Table Excel sur le bloc "par SIREN" si présent
        if not resume_par_siren.empty:
            nrows, ncols = resume_par_siren.shape
            last_col = get_column_letter(ncols)
            table_top = start_row + 1  # startrow est 0-index côté pandas, Excel est 1-index
            table_bottom = start_row + 1 + nrows  # + header
            table_ref = f"A{table_top}:{last_col}{table_bottom}"

            t = Table(displayName="T_RESUME", ref=table_ref)
            t.tableStyleInfo = TableStyleInfo(
                name="TableStyleLight9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws_r.add_table(t)

    output.seek(0)
    return output.getvalue()
